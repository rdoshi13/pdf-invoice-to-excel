from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from src.excel_writer import write_invoices
from src.models import Invoice, InvoiceItem


def make_invoice(order_date: date, source: str = "invoice.pdf") -> Invoice:
    return Invoice(
        source_path=Path(source),
        store_name="Walmart",
        order_date=order_date,
        order_number="123",
        items=[
            InvoiceItem("Apple", Decimal("2.71"), quantity="2"),
            InvoiceItem("Normal Bananas", Decimal("1.42"), quantity="1"),
        ],
        tax=Decimal("0.71"),
        total=Decimal("4.84"),
        raw_text="raw",
    )


def test_write_invoice_sheet_layout_and_formulas_for_three_participants(tmp_path):
    output_path = tmp_path / "walmart_orders.xlsx"

    added = write_invoices([make_invoice(date(2024, 5, 1))], output_path, ["Alice", "Bob", "Charlie"])

    workbook = load_workbook(output_path, data_only=False)
    worksheet = workbook[added[0]]

    assert added == ["1 May 2024"]
    assert worksheet["A1"].value == "Walmart 1 May 2024"
    assert "A1:K1" in [str(cell_range) for cell_range in worksheet.merged_cells.ranges]
    assert [worksheet.cell(row=2, column=column).value for column in range(1, 12)] == [
        "Items",
        "Qty",
        "Cost",
        "Alice",
        "Bob",
        "Charlie",
        "Involved",
        "Per Person",
        "Alice",
        "Bob",
        "Charlie",
    ]
    assert worksheet["B3"].value == "2"
    assert worksheet["A5"].value == "Walmart Tax"
    assert worksheet["G3"].value == "=SUM(D3:F3)"
    assert worksheet["H3"].value == "=IF(G3=0,0,C3/G3)"
    assert worksheet["I3"].value == "=IF(D3=1,H3,0)"
    assert worksheet["K3"].value == "=IF(F3=1,H3,0)"
    assert worksheet["A6"].value == "Total"
    assert worksheet["C6"].value == "=SUM(C3:C5)"
    assert worksheet["I6"].value == "=SUM(I3:I5)"
    assert worksheet["K6"].value == "=SUM(K3:K5)"
    assert worksheet.freeze_panes == "A3"


def test_write_invoice_sheet_layout_for_one_participant(tmp_path):
    output_path = tmp_path / "walmart_orders.xlsx"

    added = write_invoices([make_invoice(date(2024, 5, 1))], output_path, ["Alice"])

    worksheet = load_workbook(output_path, data_only=False)[added[0]]
    assert "A1:G1" in [str(cell_range) for cell_range in worksheet.merged_cells.ranges]
    assert [worksheet.cell(row=2, column=column).value for column in range(1, 8)] == [
        "Items",
        "Qty",
        "Cost",
        "Alice",
        "Involved",
        "Per Person",
        "Alice",
    ]
    assert worksheet["E3"].value == "=SUM(D3:D3)"
    assert worksheet["F3"].value == "=IF(E3=0,0,C3/E3)"
    assert worksheet["G3"].value == "=IF(D3=1,F3,0)"


def test_write_invoice_sheet_layout_for_six_participants(tmp_path):
    output_path = tmp_path / "walmart_orders.xlsx"
    participants = ["A", "B", "C", "D", "E", "F"]

    added = write_invoices([make_invoice(date(2024, 5, 1))], output_path, participants)

    worksheet = load_workbook(output_path, data_only=False)[added[0]]
    assert "A1:Q1" in [str(cell_range) for cell_range in worksheet.merged_cells.ranges]
    assert worksheet["I2"].value == "F"
    assert worksheet["J2"].value == "Involved"
    assert worksheet["K2"].value == "Per Person"
    assert worksheet["Q2"].value == "F"
    assert worksheet["J3"].value == "=SUM(D3:I3)"
    assert worksheet["K3"].value == "=IF(J3=0,0,C3/J3)"
    assert worksheet["Q3"].value == "=IF(I3=1,K3,0)"


def test_duplicate_sheet_names_append_safely(tmp_path):
    output_path = tmp_path / "walmart_orders.xlsx"

    first = write_invoices([make_invoice(date(2024, 6, 16))], output_path, ["Alice", "Bob"])
    second = write_invoices([make_invoice(date(2024, 6, 16), "other.pdf")], output_path, ["Alice", "Bob"])

    assert first == ["16 June 2024"]
    assert second == ["16 June 2024 2"]
    workbook = load_workbook(output_path)
    assert "16 June 2024" in workbook.sheetnames
    assert "16 June 2024 2" in workbook.sheetnames


def test_worksheets_are_sorted_by_date(tmp_path):
    output_path = tmp_path / "walmart_orders.xlsx"

    added = write_invoices(
        [
            make_invoice(date(2024, 8, 18)),
            make_invoice(date(2024, 5, 12)),
            make_invoice(date(2025, 2, 18)),
            make_invoice(date(2024, 4, 20)),
        ],
        output_path,
        ["Alice", "Bob"],
    )

    assert added == ["20 April 2024", "12 May 2024", "18 August 2024", "18 February 2025"]
    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["20 April 2024", "12 May 2024", "18 August 2024", "18 February 2025"]


def test_title_and_tax_row_use_invoice_store_name(tmp_path):
    output_path = tmp_path / "orders.xlsx"
    invoice = make_invoice(date(2024, 7, 10))
    invoice.store_name = "Target"

    added = write_invoices([invoice], output_path, ["Alice", "Bob"])

    worksheet = load_workbook(output_path, data_only=False)[added[0]]
    assert worksheet["A1"].value == "Target 10 July 2024"
    assert worksheet["A5"].value == "Target Tax"
