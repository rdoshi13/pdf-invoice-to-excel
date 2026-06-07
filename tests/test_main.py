from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from src import main
from src.models import Invoice, InvoiceItem


def test_cli_processes_multiple_pdfs_with_mocked_extraction(tmp_path, monkeypatch):
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    (input_dir / "b.pdf").write_text("pdf", encoding="utf-8")
    (input_dir / "a.pdf").write_text("pdf", encoding="utf-8")
    (input_dir / "notes.txt").write_text("skip", encoding="utf-8")
    output_path = tmp_path / "output" / "walmart_orders.xlsx"

    def fake_extract_text(pdf_path: Path) -> str:
        return f"raw {pdf_path.name}"

    def fake_parse_invoice_text(raw_text: str, source_path: Path) -> Invoice:
        day = 1 if source_path.name == "a.pdf" else 2
        return Invoice(
            source_path=source_path,
            order_date=date(2024, 5, day),
            order_number=source_path.stem,
            items=[InvoiceItem("Apple", Decimal("2.00"))],
            tax=Decimal("0.10"),
            total=Decimal("2.10"),
            raw_text=raw_text,
        )

    monkeypatch.setattr(main, "extract_text", fake_extract_text)
    monkeypatch.setattr(main, "parse_invoice_text", fake_parse_invoice_text)

    result = main.process_invoices(input_dir, output_path)

    assert result.found_pdf_count == 2
    assert result.parsed_count == 2
    assert result.failed_files == []
    assert result.added_sheets == ["1 May 2024", "2 May 2024"]
    assert load_workbook(output_path).sheetnames == ["1 May 2024", "2 May 2024"]
