from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from models import Invoice, InvoiceItem


GREEN_FILL = PatternFill("solid", fgColor="D9EAD3")
HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
CURRENCY_FORMAT = "0.00"
INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")
DATE_SHEET_RE = re.compile(r"^(?P<day>\d{1,2})\s+(?P<month>[A-Za-z]+)\s+(?P<year>\d{4})(?:\s+\d+)?$")


def write_invoices(invoices: list[Invoice], output_path: Path, participants: list[str]) -> list[str]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(output_path) if output_path.exists() else Workbook()

    if workbook.sheetnames == ["Sheet"] and not workbook["Sheet"].max_row > 1:
        workbook.remove(workbook["Sheet"])

    added_sheets: list[str] = []
    for invoice in sorted(invoices, key=lambda current_invoice: current_invoice.order_date):
        sheet_name = unique_sheet_name(workbook.sheetnames, date_sheet_name(invoice.order_date))
        worksheet = workbook.create_sheet(sheet_name)
        write_invoice_sheet(worksheet, invoice, participants)
        added_sheets.append(sheet_name)

    sort_worksheets_by_date(workbook)
    workbook.save(output_path)
    return added_sheets


def write_invoice_sheet(worksheet: Worksheet, invoice: Invoice, participants: list[str]) -> None:
    title = f"{invoice.store_name} {date_sheet_name(invoice.order_date)}"
    rows = list(invoice.items)
    if invoice.tax != Decimal("0"):
        rows.append(InvoiceItem(name=f"{invoice.store_name} Tax", cost=invoice.tax))

    layout = SheetLayout(participants)
    _write_title(worksheet, title, layout)
    _write_headers(worksheet, layout)
    _write_item_rows(worksheet, rows, layout)
    _write_total_row(worksheet, len(rows), layout)
    _format_sheet(worksheet, len(rows), layout)


def date_sheet_name(order_date: date) -> str:
    return f"{order_date.day} {order_date.strftime('%B')} {order_date.year}"


def sort_worksheets_by_date(workbook: Workbook) -> None:
    original_positions = {worksheet.title: index for index, worksheet in enumerate(workbook.worksheets)}

    def sort_key(worksheet: Worksheet) -> tuple[int, date, int]:
        sheet_date = _date_from_sheet_name(worksheet.title)
        if sheet_date is None:
            return (1, date.max, original_positions[worksheet.title])
        return (0, sheet_date, original_positions[worksheet.title])

    workbook._sheets = sorted(workbook._sheets, key=sort_key)


def unique_sheet_name(existing_names: list[str], desired_name: str) -> str:
    base = sanitize_sheet_name(desired_name)
    if base not in existing_names:
        return base

    counter = 2
    while True:
        suffix = f" {counter}"
        candidate = sanitize_sheet_name(f"{base[:31 - len(suffix)]}{suffix}")
        if candidate not in existing_names:
            return candidate
        counter += 1


def sanitize_sheet_name(name: str) -> str:
    sanitized = INVALID_SHEET_CHARS.sub(" ", name).strip() or "Sheet"
    return sanitized[:31]


def _date_from_sheet_name(sheet_name: str) -> date | None:
    match = DATE_SHEET_RE.match(sheet_name)
    if not match:
        return None

    try:
        return date(
            int(match.group("year")),
            _month_number(match.group("month")),
            int(match.group("day")),
        )
    except ValueError:
        return None


def _month_number(month_name: str) -> int:
    months = {
        "January": 1,
        "February": 2,
        "March": 3,
        "April": 4,
        "May": 5,
        "June": 6,
        "July": 7,
        "August": 8,
        "September": 9,
        "October": 10,
        "November": 11,
        "December": 12,
    }
    return months[month_name]


class SheetLayout:
    def __init__(self, participants: list[str]) -> None:
        self.participants = participants
        self.item_column = 1
        self.quantity_column = 2
        self.cost_column = 3
        self.first_flag_column = 4
        self.involved_column = self.first_flag_column + len(participants)
        self.per_person_column = self.involved_column + 1
        self.first_owed_column = self.per_person_column + 1
        self.last_column = self.first_owed_column + len(participants) - 1

    @property
    def flag_columns(self) -> list[int]:
        return list(range(self.first_flag_column, self.first_flag_column + len(self.participants)))

    @property
    def owed_columns(self) -> list[int]:
        return list(range(self.first_owed_column, self.first_owed_column + len(self.participants)))

    @property
    def headers(self) -> list[str]:
        return [
            "Items",
            "Qty",
            "Cost",
            *self.participants,
            "Involved",
            "Per Person",
            *self.participants,
        ]

    @property
    def title_range(self) -> str:
        return f"A1:{get_column_letter(self.last_column)}1"


def _write_title(worksheet: Worksheet, title: str, layout: SheetLayout) -> None:
    worksheet.merge_cells(layout.title_range)
    cell = worksheet["A1"]
    cell.value = title
    cell.fill = GREEN_FILL
    cell.font = Font(size=22)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 34


def _write_headers(worksheet: Worksheet, layout: SheetLayout) -> None:
    for column_index, header in enumerate(layout.headers, start=1):
        cell = worksheet.cell(row=2, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_item_rows(worksheet: Worksheet, rows: list[InvoiceItem], layout: SheetLayout) -> None:
    for row_offset, item in enumerate(rows, start=3):
        worksheet.cell(row=row_offset, column=layout.item_column, value=item.name)
        worksheet.cell(row=row_offset, column=layout.quantity_column, value=item.quantity)
        worksheet.cell(row=row_offset, column=layout.cost_column, value=float(item.cost))

        first_flag = get_column_letter(layout.flag_columns[0])
        last_flag = get_column_letter(layout.flag_columns[-1])
        involved = get_column_letter(layout.involved_column)
        per_person = get_column_letter(layout.per_person_column)
        cost = get_column_letter(layout.cost_column)
        worksheet.cell(row=row_offset, column=layout.involved_column, value=f"=SUM({first_flag}{row_offset}:{last_flag}{row_offset})")
        worksheet.cell(row=row_offset, column=layout.per_person_column, value=f"=IF({involved}{row_offset}=0,0,{cost}{row_offset}/{involved}{row_offset})")

        for flag_column, owed_column in zip(layout.flag_columns, layout.owed_columns, strict=True):
            flag_letter = get_column_letter(flag_column)
            owed_letter = get_column_letter(owed_column)
            worksheet[f"{owed_letter}{row_offset}"] = f"=IF({flag_letter}{row_offset}=1,{per_person}{row_offset},0)"


def _write_total_row(worksheet: Worksheet, item_count: int, layout: SheetLayout) -> None:
    total_row = item_count + 3
    last_item_row = total_row - 1
    cost_letter = get_column_letter(layout.cost_column)
    worksheet.cell(row=total_row, column=layout.item_column, value="Total")
    worksheet.cell(row=total_row, column=layout.cost_column, value=f"=SUM({cost_letter}3:{cost_letter}{last_item_row})")

    for owed_column in layout.owed_columns:
        owed_letter = get_column_letter(owed_column)
        worksheet[f"{owed_letter}{total_row}"] = f"=SUM({owed_letter}3:{owed_letter}{last_item_row})"

    for column in range(1, layout.last_column + 1):
        worksheet.cell(row=total_row, column=column).font = Font(bold=True)


def _format_sheet(worksheet: Worksheet, item_count: int, layout: SheetLayout) -> None:
    worksheet.freeze_panes = "A3"

    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions[get_column_letter(layout.quantity_column)].width = 10
    for column_index in range(3, layout.last_column + 1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = 12

    final_row = item_count + 3
    for row in range(3, final_row + 1):
        money_columns = [layout.cost_column, layout.per_person_column, *layout.owed_columns]
        for column in money_columns:
            worksheet.cell(row=row, column=column).number_format = CURRENCY_FORMAT
        for column in range(2, layout.last_column + 1):
            worksheet.cell(row=row, column=column).alignment = Alignment(horizontal="right")

    for column_index in range(1, layout.last_column + 1):
        column_letter = get_column_letter(column_index)
        worksheet[f"{column_letter}1"].fill = GREEN_FILL
